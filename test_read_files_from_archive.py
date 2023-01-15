import os
import zipfile
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from create_archive import create_archive
import pytest


path_files = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')
path_resources = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
path_zip = os.path.join(path_resources, "qa_guru_3.6.zip")


def test_contains_csv_in_archive():
    create_archive(path_files, path_zip)
    with zipfile.ZipFile(path_zip) as zf:
        cf = zf.extract("Test_case_csv.csv")

        with open(cf) as csvfile:
            list_csv = []
            csvfile = csv.reader(csvfile)
            for r in csvfile:
                text_in_csv = '; '.join(r)
                list_csv.append(text_in_csv)
            assert list_csv[1] == '2021; Level 1; 99999; All', f'Текст в в файле .csv не совпадает\n' \
                                                                 f'Ожидаемый результат: {"Year,Industry_aggregation_NZSIOC,Industry_code_NZSIOC"}\n' \
                                                                 f'Фактический результат: {list_csv[1]}\n'
    os.remove(cf)


def test_contains_pdf_in_archive():
    create_archive(path_files, path_zip)
    with zipfile.ZipFile(path_zip) as zf:
        cf = zf.extract("Test.pdf")

        reader = PdfReader(cf)
        page = reader.pages[0]
        text = page.extract_text()
        assert 'T E S T' in text, f'Текст: {"T E S T"} не содержится в {text}'
        os.remove(cf)


def test_contains_xlsx_in_archive():
    create_archive(path_files, path_zip)
    with zipfile.ZipFile(path_zip) as zf:
        cf = zf.extract("Test_case_xlsx.xlsx")

        workbook = load_workbook(cf)
        sheet = workbook.active
        name_user = sheet.cell(row=2, column=2).value
        assert name_user == 'Добавление в поле "Имя"', 'Значение колонки в таблице не совпадает!'
        sheet = workbook.close()
    os.remove(cf)